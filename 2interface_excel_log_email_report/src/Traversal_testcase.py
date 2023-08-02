# /usr/bin/env python
# coding=utf-8
import openpyxl
from src.interfaceclass import *
from commons.log import log


class run_case():
    def runcase(self, test_case, save_case):
        logg = log()
        # try:
        #     wp = openpyxl.load_workbook(test_case)
        # except Exception as e:
        #     # print("打开用例失败！")
        #     logg.error("打开测试用例出错,测试用例路径：",test_case)
        wp = openpyxl.load_workbook(test_case)
        sheet = wp.get_sheet_by_name('TestCase')
        corr_dict = {}

        for i in range(2, sheet.max_row + 1):
            if sheet.cell(row=i, column=10).value.replace("\n", "").replace("\r", "") == "Yes":
                # 如果用例中该字段为Yes则不执行跳出
                continue
            params = sheet.cell(row=i, column=7).value.replace("\n", "").replace("\r", "")
            # print("params==",params)
            url1 = sheet.cell(row=i, column=3).value.replace("\n", "").replace("\r", "")
            url2 = sheet.cell(row=i, column=4).value.replace("\n", "").replace("\r", "")
            url = url1 + url2
            requ_method = sheet.cell(row=i, column=5).value.replace("\n", "").replace("\r", "")
            types = sheet.cell(row=i, column=6).value.replace("\n", "").replace("\r", "")
            chinkoption = sheet.cell(row=i, column=8).value.replace("\n", "").replace("\r", "")
            headers = {}

            # 替换请求参数
            for keyword in corr_dict:
                # print("keyword=",keyword)   #打印keyword= ${date}
                if params.find(keyword) > 0:
                    # print("params=",params)  #打印params= {"key":"e711bc6362b3179f5a28de7fd3ee4ace","date":"${date}"}
                    # print(corr_dict[keyword]) # 2023-08-02
                    params = params.replace(keyword, str(corr_dict[keyword]))
                    # print('===',params)  # {"key":"e711bc6362b3179f5a28de7fd3ee4ace","date":"2023-08-02"}

            params = eval(params)
            q = InterfaceTest()
            # q.testsearch(url,params,headers,requ_method,chinkoption,types,sheet,i)
            res = q.testsearch(url, params, headers, requ_method, chinkoption, types, sheet, i, logg)

            if sheet.cell(row=i, column=9).value != None:
                corr = sheet.cell(row=i, column=9).value.replace('\n', '').replace('\r', '')
                corr = corr.replace('\n', '').replace('\r', '').split(";")
                # print('111',corr)  # ['${date}=[result][yangli]']
                for j in range(len(corr)):
                    param = corr[j].split("=")
                    # print('222',param)  # ['${date}', '[result][yangli]']
                    for key in param[1][1:-1].split("]["):
                        # print('222-1',param[1][1:-1])  # result][yangli
                        # print('222-2',param[1][1:-1].split("][")) # ['result', 'yangli']
                        temp = res[key]
                        # print('res[key]',res[key]) # {'id': '4804', 'yangli': '2023-08-02', 'yinli': '癸卯(兔)年六月十六',
                        # 'wuxing': '长流水 收执位', 'chongsha': '冲狗(丙戍)煞南', 'baiji': '壬不汲水更难提防 辰不哭泣必主重丧', 'jishen': '天马
                        # 普护', 'yi': '祭祀 冠笄 作灶 交易 纳财 栽种 结网 纳畜 牧养 进人口', 'xiongshen': '五虚 白虎', 'ji': '开渠 造船 安床 安葬 破土 出行
                        # 修坟 掘井 开市 开生坟'}
                        res = temp
                        print('444',res) # 遍历result：{'id': '4804', 'yangli': '2023-08-02', 'yinli': '癸卯(兔)年六月十六', 'wuxing': '长流水 收执位', 'chongsha': '冲狗(丙戍)煞南', 'baiji': '壬不汲水更难提防 辰不哭泣必主重丧', 'jishen': '天马 普护', 'yi': '祭祀 冠笄 作灶 交易 纳财 栽种 结网 纳畜 牧养 进人口', 'xiongshen': '五虚 白虎', 'ji': '开渠 造船 安床 安葬 破土 出行 修坟 掘井 开市 开生坟'}
                                         # 遍历yangli：2023-08-02
                corr_dict[param[0]] = res
                print('555',corr_dict)  # 将遍历的最后一个元素的值赋值给了corr_dict: {'${date}': '2023-08-02'}

        # 关联参数的拆分

        wp.save(save_case)


# test_case = "D:/pycharm workspace/framework/testcase/laohuanglitestcase5.xlsx"
# save_case = "D:/pycharm workspace/framework/report/guo.xlsx"
# run = run_case()
# run.runcase(test_case,save_case)


'''
1关联的格式
2两个关联的参数用分号；隔开
3关联参数的拆分
多次for循环进行拆分
一次是根据；进行拆分
二次是根据=进行拆分
三次是根据][进行拆分
得到key，根据key去取value值
定义一个空字典把对应的值放进去
4关联参数去替换请求中的参数
for循环去遍历key，如果找到就替换为key对应的value
'''
